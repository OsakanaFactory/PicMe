from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# スタイル定義
header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', start_color='4472C4')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_sheet(wb, title, test_cases):
    if title == "サマリ":
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)

    headers = ['テストID', 'カテゴリ', '機能', 'テスト名', 'テスト種別', '前提条件', 'テスト手順', '期待結果', 'テスト結果', '実施日', '備考']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row_num, row_data in enumerate(test_cases, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border

    column_widths = [12, 12, 12, 25, 10, 25, 55, 30, 12, 12, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for row in range(2, len(test_cases) + 2):
        ws.row_dimensions[row].height = 70

    return ws

# ========== 認証テスト ==========
auth_tests = [
    ["AUTH-001", "認証", "サインアップ", "正常なサインアップ", "正常系", "未登録ユーザー",
     "1. /api/auth/signup にPOST\n2. username: 「test_user」\n3. email: 「test@example.com」\n4. password: 「Password123!」",
     "status: 201\naccessToken/refreshToken返却", "", "", ""],
    ["AUTH-002", "認証", "サインアップ", "既存ユーザー名", "異常系", "username「test_user」が存在",
     "1. /api/auth/signup にPOST\n2. username: 「test_user」（既存）\n3. email: 「new@example.com」\n4. password: 「Password123!」",
     "status: 409\nerror: USERNAME_ALREADY_EXISTS", "", "", ""],
    ["AUTH-003", "認証", "サインアップ", "既存メール", "異常系", "email「test@example.com」が存在",
     "1. /api/auth/signup にPOST\n2. username: 「new_user」\n3. email: 「test@example.com」（既存）\n4. password: 「Password123!」",
     "status: 409\nerror: EMAIL_ALREADY_EXISTS", "", "", ""],
    ["AUTH-004", "認証", "サインアップ", "無効なメール形式", "異常系", "なし",
     "1. /api/auth/signup にPOST\n2. email: 「invalid-email」（@なし）",
     "status: 400\nバリデーションエラー", "", "", ""],
    ["AUTH-005", "認証", "サインアップ", "パスワード7文字", "境界値", "なし",
     "1. /api/auth/signup にPOST\n2. password: 「Pass1!a」（7文字）",
     "status: 400\nパスワードは8文字以上", "", "", ""],
    ["AUTH-006", "認証", "サインアップ", "パスワード8文字", "境界値", "なし",
     "1. /api/auth/signup にPOST\n2. username: 「pass_test」\n3. email: 「pass@example.com」\n4. password: 「Pass123!」（8文字）",
     "status: 201\n成功", "", "", ""],
    ["AUTH-007", "認証", "サインアップ", "ユーザー名2文字", "境界値", "なし",
     "1. /api/auth/signup にPOST\n2. username: 「ab」（2文字）",
     "status: 400\nユーザー名は3文字以上", "", "", ""],
    ["AUTH-008", "認証", "サインアップ", "ユーザー名3文字", "境界値", "なし",
     "1. /api/auth/signup にPOST\n2. username: 「abc」（3文字）\n3. email: 「abc@example.com」\n4. password: 「Password123!」",
     "status: 201\n成功", "", "", ""],
    ["AUTH-009", "認証", "サインアップ", "ユーザー名20文字", "境界値", "なし",
     "1. /api/auth/signup にPOST\n2. username: 「abcdefghij1234567890」（20文字）\n3. email: 「max@example.com」\n4. password: 「Password123!」",
     "status: 201\n成功", "", "", ""],
    ["AUTH-010", "認証", "サインアップ", "ユーザー名21文字", "境界値", "なし",
     "1. /api/auth/signup にPOST\n2. username: 「abcdefghij12345678901」（21文字）",
     "status: 400\nユーザー名は20文字以内", "", "", ""],
    ["AUTH-011", "認証", "ログイン", "正常なログイン", "正常系", "登録済みユーザー",
     "1. /api/auth/login にPOST\n2. email: 「test@example.com」\n3. password: 「Password123!」",
     "status: 200\naccessToken/refreshToken返却", "", "", ""],
    ["AUTH-012", "認証", "ログイン", "無効なパスワード", "異常系", "登録済みユーザー",
     "1. /api/auth/login にPOST\n2. email: 「test@example.com」\n3. password: 「WrongPass123!」",
     "status: 401\nerror: INVALID_CREDENTIALS", "", "", ""],
    ["AUTH-013", "認証", "ログイン", "存在しないメール", "異常系", "なし",
     "1. /api/auth/login にPOST\n2. email: 「notexist@example.com」\n3. password: 「Password123!」",
     "status: 401\nerror: INVALID_CREDENTIALS", "", "", ""],
    ["AUTH-014", "認証", "ログイン", "空のメール", "異常系", "なし",
     "1. /api/auth/login にPOST\n2. email: 「」（空）\n3. password: 「Password123!」",
     "status: 400\nバリデーションエラー", "", "", ""],
    ["AUTH-015", "認証", "ログイン", "空のパスワード", "異常系", "なし",
     "1. /api/auth/login にPOST\n2. email: 「test@example.com」\n3. password: 「」（空）",
     "status: 400\nバリデーションエラー", "", "", ""],
    ["AUTH-016", "認証", "トークン更新", "正常なトークン更新", "正常系", "有効なrefreshToken",
     "1. /api/auth/refresh にPOST\n2. refreshToken: 「{取得したrefreshToken}」",
     "status: 200\n新accessToken返却", "", "", ""],
    ["AUTH-017", "認証", "トークン更新", "無効なrefreshToken", "異常系", "なし",
     "1. /api/auth/refresh にPOST\n2. refreshToken: 「invalid_token_12345」",
     "status: 401\nerror: INVALID_TOKEN", "", "", ""],
    ["AUTH-018", "認証", "認証情報取得", "認証済みユーザー", "正常系", "ログイン済み",
     "1. /api/auth/me にGET\n2. Header: Authorization: Bearer {accessToken}",
     "status: 200\nユーザー情報返却", "", "", ""],
    ["AUTH-019", "認証", "認証情報取得", "未認証", "異常系", "未ログイン",
     "1. /api/auth/me にGET\n2. Authorizationヘッダーなし",
     "status: 401\nerror: UNAUTHORIZED", "", "", ""],
    ["AUTH-020", "認証", "認証情報取得", "無効なトークン", "異常系", "なし",
     "1. /api/auth/me にGET\n2. Header: Authorization: Bearer invalid_token",
     "status: 401\nerror: INVALID_TOKEN", "", "", ""],
    ["AUTH-021", "認証", "ログアウト", "正常なログアウト", "正常系", "ログイン済み",
     "1. /api/auth/logout にPOST\n2. Header: Authorization: Bearer {accessToken}",
     "status: 200\n成功メッセージ", "", "", ""],
]

# ========== プロフィールテスト ==========
profile_tests = [
    ["PROF-001", "プロフィール", "取得", "自分のプロフィール取得", "正常系", "ログイン済み",
     "1. /api/profile にGET\n2. Header: Authorization: Bearer {accessToken}",
     "status: 200\nプロフィール情報返却", "", "", ""],
    ["PROF-002", "プロフィール", "取得", "未認証", "異常系", "未ログイン",
     "1. /api/profile にGET\n2. Authorizationヘッダーなし",
     "status: 401", "", "", ""],
    ["PROF-003", "プロフィール", "更新", "表示名の更新", "正常系", "ログイン済み",
     "1. /api/profile にPUT\n2. displayName: 「新しい表示名」",
     "status: 200\n更新されたプロフィール", "", "", ""],
    ["PROF-004", "プロフィール", "更新", "自己紹介の更新", "正常系", "ログイン済み",
     "1. /api/profile にPUT\n2. bio: 「新しい自己紹介文です。」",
     "status: 200", "", "", ""],
    ["PROF-005", "プロフィール", "更新", "全フィールド更新", "正常系", "ログイン済み",
     "1. /api/profile にPUT\n2. displayName: 「テストユーザー」\n3. bio: 「自己紹介」\n4. avatarUrl: 「https://example.com/avatar.jpg」",
     "status: 200\n全フィールド更新確認", "", "", ""],
    ["PROF-006", "プロフィール", "更新", "表示名32文字", "境界値", "ログイン済み",
     "1. /api/profile にPUT\n2. displayName: 「あいうえおかきくけこさしすせそたちつてとなにぬねのはひふへほまみ」（32文字）",
     "status: 200\n成功", "", "", ""],
    ["PROF-007", "プロフィール", "更新", "表示名33文字", "境界値", "ログイン済み",
     "1. /api/profile にPUT\n2. displayName: 「あいうえおかきくけこさしすせそたちつてとなにぬねのはひふへほまみむ」（33文字）",
     "status: 400\n32文字以内", "", "", ""],
    ["PROF-008", "プロフィール", "更新", "自己紹介500文字", "境界値", "ログイン済み",
     "1. /api/profile にPUT\n2. bio: 「あ」×500（500文字）",
     "status: 200\n成功", "", "", ""],
    ["PROF-009", "プロフィール", "更新", "自己紹介501文字", "境界値", "ログイン済み",
     "1. /api/profile にPUT\n2. bio: 「あ」×501（501文字）",
     "status: 400\n500文字以内", "", "", ""],
]

# ========== 作品テスト ==========
artwork_tests = [
    ["ART-001", "作品", "一覧取得", "自分の作品一覧", "正常系", "ログイン済み/作品あり",
     "1. /api/artworks にGET\n2. Header: Authorization: Bearer {accessToken}",
     "status: 200\n作品一覧返却", "", "", ""],
    ["ART-002", "作品", "一覧取得", "作品なし", "正常系", "ログイン済み/作品なし",
     "1. /api/artworks にGET",
     "status: 200\n空配列[]", "", "", ""],
    ["ART-003", "作品", "作成", "新規作品作成", "正常系", "ログイン済み",
     "1. /api/artworks にPOST\n2. title: 「テスト作品」\n3. imageUrl: 「https://example.com/image.jpg」\n4. description: 「作品の説明」",
     "status: 201\n作成された作品", "", "", ""],
    ["ART-004", "作品", "作成", "タイトルなし", "異常系", "ログイン済み",
     "1. /api/artworks にPOST\n2. title: 「」（空）\n3. imageUrl: 「https://example.com/image.jpg」",
     "status: 400\ntitleは必須", "", "", ""],
    ["ART-005", "作品", "作成", "画像URLなし", "異常系", "ログイン済み",
     "1. /api/artworks にPOST\n2. title: 「テスト」\n3. imageUrl: 「」（空）",
     "status: 400\nimageUrlは必須", "", "", ""],
    ["ART-006", "作品", "作成", "タイトル200文字", "境界値", "ログイン済み",
     "1. /api/artworks にPOST\n2. title: 「あ」×200（200文字）\n3. imageUrl: 「https://example.com/image.jpg」",
     "status: 201\n成功", "", "", ""],
    ["ART-007", "作品", "作成", "タイトル201文字", "境界値", "ログイン済み",
     "1. /api/artworks にPOST\n2. title: 「あ」×201（201文字）",
     "status: 400\n200文字以内", "", "", ""],
    ["ART-008", "作品", "取得", "作品詳細取得", "正常系", "ログイン済み/作品あり",
     "1. /api/artworks/{id} にGET\n2. id: 「1」（自分の作品ID）",
     "status: 200\n作品詳細", "", "", ""],
    ["ART-009", "作品", "取得", "存在しない作品", "異常系", "ログイン済み",
     "1. /api/artworks/{id} にGET\n2. id: 「99999」",
     "status: 404\nNOT_FOUND", "", "", ""],
    ["ART-010", "作品", "取得", "他ユーザーの作品", "異常系", "ログイン済み",
     "1. /api/artworks/{id} にGET\n2. id: 「{他ユーザーの作品ID}」",
     "status: 403\nFORBIDDEN", "", "", ""],
    ["ART-011", "作品", "更新", "タイトル更新", "正常系", "ログイン済み/作品あり",
     "1. /api/artworks/{id} にPUT\n2. id: 「1」\n3. title: 「更新後タイトル」",
     "status: 200\n更新された作品", "", "", ""],
    ["ART-012", "作品", "更新", "他ユーザーの作品", "異常系", "ログイン済み",
     "1. /api/artworks/{id} にPUT\n2. id: 「{他ユーザーの作品ID}」\n3. title: 「更新」",
     "status: 403\nFORBIDDEN", "", "", ""],
    ["ART-013", "作品", "削除", "作品削除", "正常系", "ログイン済み/作品あり",
     "1. /api/artworks/{id} にDELETE\n2. id: 「1」",
     "status: 200 or 204", "", "", ""],
    ["ART-014", "作品", "削除", "他ユーザーの作品", "異常系", "ログイン済み",
     "1. /api/artworks/{id} にDELETE\n2. id: 「{他ユーザーの作品ID}」",
     "status: 403\nFORBIDDEN", "", "", ""],
    ["ART-015", "作品", "並び替え", "並び替え", "正常系", "ログイン済み/複数作品",
     "1. /api/artworks/reorder にPUT\n2. orderedIds: [3, 1, 2]",
     "status: 200\n並び順更新", "", "", ""],
    ["ART-016", "作品", "プラン制限", "FREE: 5作品目", "プラン制限", "FREEプラン/4作品あり",
     "1. /api/artworks にPOST\n2. title: 「5個目の作品」\n3. imageUrl: 「https://example.com/5.jpg」",
     "status: 201\n成功", "", "", ""],
    ["ART-017", "作品", "プラン制限", "FREE: 6作品目", "プラン制限", "FREEプラン/5作品あり",
     "1. /api/artworks にPOST\n2. title: 「6個目の作品」\n3. imageUrl: 「https://example.com/6.jpg」",
     "status: 403\nLIMIT_EXCEEDED", "", "", ""],
]

# ========== SNSリンクテスト ==========
social_tests = [
    ["SOCIAL-001", "SNSリンク", "一覧取得", "一覧取得", "正常系", "ログイン済み",
     "1. /api/social-links にGET",
     "status: 200\nリンク一覧", "", "", ""],
    ["SOCIAL-002", "SNSリンク", "作成", "新規作成", "正常系", "ログイン済み",
     "1. /api/social-links にPOST\n2. platform: 「Twitter」\n3. url: 「https://twitter.com/testuser」",
     "status: 201\n作成されたリンク", "", "", ""],
    ["SOCIAL-003", "SNSリンク", "作成", "URLなし", "異常系", "ログイン済み",
     "1. /api/social-links にPOST\n2. platform: 「Twitter」\n3. url: 「」（空）",
     "status: 400\nurlは必須", "", "", ""],
    ["SOCIAL-004", "SNSリンク", "作成", "無効なURL", "異常系", "ログイン済み",
     "1. /api/social-links にPOST\n2. platform: 「Twitter」\n3. url: 「invalid-url」",
     "status: 400\nURL形式エラー", "", "", ""],
    ["SOCIAL-005", "SNSリンク", "更新", "URL更新", "正常系", "ログイン済み/リンクあり",
     "1. /api/social-links/{id} にPUT\n2. id: 「1」\n3. url: 「https://twitter.com/newuser」",
     "status: 200", "", "", ""],
    ["SOCIAL-006", "SNSリンク", "削除", "リンク削除", "正常系", "ログイン済み/リンクあり",
     "1. /api/social-links/{id} にDELETE\n2. id: 「1」",
     "status: 200 or 204", "", "", ""],
    ["SOCIAL-007", "SNSリンク", "並び替え", "並び替え", "正常系", "複数リンクあり",
     "1. /api/social-links/reorder にPUT\n2. orderedIds: [2, 1]",
     "status: 200", "", "", ""],
    ["SOCIAL-008", "SNSリンク", "プラン制限", "FREE: 2個目", "プラン制限", "FREEプラン/1リンクあり",
     "1. /api/social-links にPOST\n2. platform: 「Instagram」\n3. url: 「https://instagram.com/test」",
     "status: 201\n成功", "", "", ""],
    ["SOCIAL-009", "SNSリンク", "プラン制限", "FREE: 3個目", "プラン制限", "FREEプラン/2リンクあり",
     "1. /api/social-links にPOST\n2. platform: 「YouTube」\n3. url: 「https://youtube.com/@test」",
     "status: 403\nLIMIT_EXCEEDED", "", "", ""],
]

# ========== 公開ページテスト ==========
public_tests = [
    ["PUBLIC-001", "公開ページ", "取得", "公開ページ表示", "正常系", "公開プロフィールあり",
     "1. /api/public/{username} にGET\n2. username: 「test_user」",
     "status: 200\n公開情報", "", "", ""],
    ["PUBLIC-002", "公開ページ", "取得", "存在しないユーザー", "異常系", "なし",
     "1. /api/public/{username} にGET\n2. username: 「nonexistent_user」",
     "status: 404\nUSER_NOT_FOUND", "", "", ""],
    ["PUBLIC-003", "公開ページ", "取得", "作品一覧含む", "正常系", "公開作品あり",
     "1. /api/public/{username} にGET\n2. username: 「test_user」",
     "artworks配列に作品含む", "", "", ""],
    ["PUBLIC-004", "公開ページ", "取得", "SNSリンク含む", "正常系", "公開リンクあり",
     "1. /api/public/{username} にGET\n2. username: 「test_user」",
     "socialLinks配列にリンク含む", "", "", ""],
    ["PUBLIC-005", "公開ページ", "フィルタ", "カテゴリでフィルタ", "正常系", "カテゴリ紐付けあり",
     "1. /api/public/{username}?categoryId={id} にGET\n2. username: 「test_user」\n3. categoryId: 「1」",
     "フィルタされた作品のみ返却", "", "", ""],
    ["PUBLIC-006", "公開ページ", "フィルタ", "タグでフィルタ", "正常系", "タグ紐付けあり",
     "1. /api/public/{username}?tagId={id} にGET\n2. username: 「test_user」\n3. tagId: 「1」",
     "フィルタされた作品のみ返却", "", "", ""],
    ["PUBLIC-007", "公開ページ", "投稿", "公開投稿一覧", "正常系", "公開投稿あり",
     "1. /api/public/{username} にGET\n2. username: 「test_user」",
     "posts配列に投稿含む", "", "", ""],
    ["PUBLIC-008", "公開ページ", "投稿", "非公開投稿除外", "正常系", "非公開投稿あり",
     "1. /api/public/{username} にGET\n2. username: 「test_user」",
     "非公開投稿は含まない", "", "", ""],
]

# ========== 投稿テスト ==========
post_tests = [
    ["POST-001", "投稿", "一覧取得", "投稿一覧", "正常系", "ログイン済み",
     "1. /api/posts にGET",
     "status: 200\n投稿一覧", "", "", ""],
    ["POST-002", "投稿", "作成", "新規投稿", "正常系", "ログイン済み",
     "1. /api/posts にPOST\n2. title: 「テスト投稿」\n3. content: 「投稿の内容です。」",
     "status: 201\n作成された投稿", "", "", ""],
    ["POST-003", "投稿", "作成", "タイトルなし", "異常系", "ログイン済み",
     "1. /api/posts にPOST\n2. title: 「」（空）\n3. content: 「内容」",
     "status: 400\ntitleは必須", "", "", ""],
    ["POST-004", "投稿", "作成", "コンテンツなし", "異常系", "ログイン済み",
     "1. /api/posts にPOST\n2. title: 「タイトル」\n3. content: 「」（空）",
     "status: 400\ncontentは必須", "", "", ""],
    ["POST-005", "投稿", "作成", "タイトル200文字", "境界値", "ログイン済み",
     "1. /api/posts にPOST\n2. title: 「あ」×200（200文字）\n3. content: 「内容」",
     "status: 201\n成功", "", "", ""],
    ["POST-006", "投稿", "作成", "タイトル201文字", "境界値", "ログイン済み",
     "1. /api/posts にPOST\n2. title: 「あ」×201（201文字）",
     "status: 400\n200文字以内", "", "", ""],
    ["POST-007", "投稿", "更新", "投稿更新", "正常系", "ログイン済み/投稿あり",
     "1. /api/posts/{id} にPUT\n2. id: 「1」\n3. title: 「更新後タイトル」",
     "status: 200", "", "", ""],
    ["POST-008", "投稿", "削除", "投稿削除", "正常系", "ログイン済み/投稿あり",
     "1. /api/posts/{id} にDELETE\n2. id: 「1」",
     "status: 200 or 204", "", "", ""],
    ["POST-009", "投稿", "公開設定", "公開設定", "正常系", "ログイン済み",
     "1. /api/posts/{id} にPUT\n2. id: 「1」\n3. published: true",
     "status: 200\n公開状態", "", "", ""],
    ["POST-010", "投稿", "プラン制限", "FREE: 1投稿目", "プラン制限", "FREEプラン/0投稿",
     "1. /api/posts にPOST\n2. title: 「1個目の投稿」\n3. content: 「内容」",
     "status: 201\n成功", "", "", ""],
    ["POST-011", "投稿", "プラン制限", "FREE: 2投稿目", "プラン制限", "FREEプラン/1投稿あり",
     "1. /api/posts にPOST\n2. title: 「2個目の投稿」\n3. content: 「内容」",
     "status: 403\nLIMIT_EXCEEDED", "", "", ""],
]

# ========== テーマテスト ==========
theme_tests = [
    ["THEME-001", "テーマ", "取得", "テーマ設定取得", "正常系", "ログイン済み",
     "1. /api/profile にGET",
     "status: 200\nテーマ設定含む", "", "", ""],
    ["THEME-002", "テーマ", "更新", "primaryColor更新", "正常系", "ログイン済み",
     "1. /api/profile にPUT\n2. colorPrimary: 「#FF0000」",
     "status: 200\n色更新", "", "", ""],
    ["THEME-003", "テーマ", "更新", "無効なカラーコード", "異常系", "ログイン済み",
     "1. /api/profile にPUT\n2. colorPrimary: 「invalid」",
     "status: 400\n無効な色形式", "", "", ""],
    ["THEME-004", "テーマ", "更新", "layout更新", "正常系", "ログイン済み",
     "1. /api/profile にPUT\n2. layout: 「GRID」",
     "status: 200", "", "", ""],
]

# ========== カテゴリ・タグテスト ==========
category_tag_tests = [
    ["CAT-001", "カテゴリ", "一覧取得", "カテゴリ一覧", "正常系", "ログイン済み（PRO以上）",
     "1. /api/categories にGET",
     "status: 200\nカテゴリ一覧", "", "", ""],
    ["CAT-002", "カテゴリ", "作成", "新規カテゴリ", "正常系", "ログイン済み（PRO以上）",
     "1. /api/categories にPOST\n2. name: 「イラスト」",
     "status: 201\n作成されたカテゴリ", "", "", ""],
    ["CAT-003", "カテゴリ", "作成", "名前なし", "異常系", "ログイン済み",
     "1. /api/categories にPOST\n2. name: 「」（空）",
     "status: 400\nnameは必須", "", "", ""],
    ["CAT-004", "カテゴリ", "作成", "重複名", "異常系", "「イラスト」が存在",
     "1. /api/categories にPOST\n2. name: 「イラスト」",
     "status: 409 or 400\n重複エラー", "", "", ""],
    ["CAT-005", "カテゴリ", "作成", "名前50文字", "境界値", "ログイン済み（PRO以上）",
     "1. /api/categories にPOST\n2. name: 「あ」×50（50文字）",
     "status: 201\n成功", "", "", ""],
    ["CAT-006", "カテゴリ", "作成", "名前51文字", "境界値", "ログイン済み",
     "1. /api/categories にPOST\n2. name: 「あ」×51（51文字）",
     "status: 400\n50文字以内", "", "", ""],
    ["CAT-007", "カテゴリ", "更新", "カテゴリ更新", "正常系", "カテゴリあり",
     "1. /api/categories/{id} にPUT\n2. id: 「1」\n3. name: 「マンガ」",
     "status: 200", "", "", ""],
    ["CAT-008", "カテゴリ", "削除", "カテゴリ削除", "正常系", "カテゴリあり（未使用）",
     "1. /api/categories/{id} にDELETE\n2. id: 「1」",
     "status: 200 or 204", "", "", ""],
    ["CAT-009", "カテゴリ", "プラン制限", "FREE: 使用不可", "プラン制限", "FREEプラン",
     "1. /api/categories にPOST\n2. name: 「テスト」",
     "status: 403\nPLAN_REQUIRED", "", "", ""],
    ["CAT-010", "カテゴリ", "作品紐付け", "作品にカテゴリ紐付け", "正常系", "作品/カテゴリあり",
     "1. /api/artworks/{id} にPUT\n2. id: 「1」\n3. categoryId: 「1」",
     "status: 200\nカテゴリ紐付け", "", "", ""],
    ["TAG-001", "タグ", "一覧取得", "タグ一覧", "正常系", "ログイン済み（PRO以上）",
     "1. /api/tags にGET",
     "status: 200\nタグ一覧", "", "", ""],
    ["TAG-002", "タグ", "作成", "新規タグ", "正常系", "ログイン済み（PRO以上）",
     "1. /api/tags にPOST\n2. name: 「オリジナル」",
     "status: 201\n作成されたタグ", "", "", ""],
    ["TAG-003", "タグ", "作成", "名前なし", "異常系", "ログイン済み",
     "1. /api/tags にPOST\n2. name: 「」（空）",
     "status: 400\nnameは必須", "", "", ""],
    ["TAG-004", "タグ", "作成", "名前30文字", "境界値", "ログイン済み（PRO以上）",
     "1. /api/tags にPOST\n2. name: 「あ」×30（30文字）",
     "status: 201\n成功", "", "", ""],
    ["TAG-005", "タグ", "作成", "名前31文字", "境界値", "ログイン済み",
     "1. /api/tags にPOST\n2. name: 「あ」×31（31文字）",
     "status: 400\n30文字以内", "", "", ""],
    ["TAG-006", "タグ", "削除", "タグ削除", "正常系", "タグあり（未使用）",
     "1. /api/tags/{id} にDELETE\n2. id: 「1」",
     "status: 200 or 204", "", "", ""],
    ["TAG-007", "タグ", "作品紐付け", "作品にタグ紐付け", "正常系", "作品/タグあり",
     "1. /api/artworks/{id} にPUT\n2. id: 「1」\n3. tagIds: [1, 2]",
     "status: 200\nタグ紐付け", "", "", ""],
    ["TAG-008", "タグ", "プラン制限", "FREE: 使用不可", "プラン制限", "FREEプラン",
     "1. /api/tags にPOST\n2. name: 「テスト」",
     "status: 403\nPLAN_REQUIRED", "", "", ""],
]

# ========== UIテスト ==========
ui_tests = [
    ["UI-001", "UI", "ログイン画面", "ログイン画面表示", "正常系", "なし",
     "1. ブラウザで /login にアクセス",
     "ログインフォーム表示", "", "", ""],
    ["UI-002", "UI", "サインアップ画面", "サインアップ画面表示", "正常系", "なし",
     "1. ブラウザで /signup にアクセス",
     "サインアップフォーム表示", "", "", ""],
    ["UI-003", "UI", "ダッシュボード", "ダッシュボード表示", "正常系", "ログイン済み",
     "1. ブラウザで /dashboard にアクセス",
     "統計情報表示", "", "", ""],
    ["UI-004", "UI", "ダッシュボード", "未認証リダイレクト", "異常系", "未ログイン",
     "1. ブラウザで /dashboard にアクセス（未ログイン）",
     "/login にリダイレクト", "", "", ""],
    ["UI-005", "UI", "サイドバー", "サイドバー表示", "正常系", "ログイン済み",
     "1. ダッシュボードでサイドバー確認\n2. 各メニュー項目をクリック",
     "全メニュー項目表示・遷移", "", "", ""],
    ["UI-006", "UI", "レスポンシブ", "モバイル表示", "正常系", "なし",
     "1. 開発者ツールでモバイルサイズ（375px）\n2. 各画面を確認",
     "レスポンシブ対応", "", "", ""],
    ["UI-007", "UI", "公開ページ", "公開ページ表示", "正常系", "公開プロフィールあり",
     "1. ブラウザで /{username} にアクセス\n2. username: 「test_user」",
     "公開ページ正常表示", "", "", ""],
]

# ========== サマリシート ==========
summary_data = [
    ["Phase1 - 認証", "21件", "", "", "", "", "", "", "", "", ""],
    ["Phase1 - プロフィール", "9件", "", "", "", "", "", "", "", "", ""],
    ["Phase1 - 作品", "17件", "", "", "", "", "", "", "", "", ""],
    ["Phase1 - SNSリンク", "9件", "", "", "", "", "", "", "", "", ""],
    ["Phase1 - 公開ページ", "8件", "", "", "", "", "", "", "", "", ""],
    ["Phase1 - UI", "7件", "", "", "", "", "", "", "", "", ""],
    ["Phase2 - 投稿", "11件", "", "", "", "", "", "", "", "", ""],
    ["Phase2 - テーマ", "4件", "", "", "", "", "", "", "", "", ""],
    ["Phase2 - カテゴリ・タグ", "18件", "", "", "", "", "", "", "", "", ""],
    ["合計", "104件", "", "", "", "", "", "", "", "", ""],
]

# シート作成
create_sheet(wb, "サマリ", summary_data)
create_sheet(wb, "認証", auth_tests)
create_sheet(wb, "プロフィール", profile_tests)
create_sheet(wb, "作品", artwork_tests)
create_sheet(wb, "SNSリンク", social_tests)
create_sheet(wb, "公開ページ", public_tests)
create_sheet(wb, "投稿", post_tests)
create_sheet(wb, "テーマ", theme_tests)
create_sheet(wb, "カテゴリ・タグ", category_tag_tests)
create_sheet(wb, "UI", ui_tests)

# 保存
output_path = '/mnt/c/OsakanaFactory/workspace/PicMe/docs/test/PicMe_テスト仕様書_v2.1.xlsx'
wb.save(output_path)
print(f"Saved to: {output_path}")
print(f"Total test cases: {len(auth_tests) + len(profile_tests) + len(artwork_tests) + len(social_tests) + len(public_tests) + len(post_tests) + len(theme_tests) + len(category_tag_tests) + len(ui_tests)}")
